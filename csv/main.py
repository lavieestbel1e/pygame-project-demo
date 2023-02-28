from openpyxl import load_workbook
import csv
lst = []
workbook_values = load_workbook(filename='data.xlsx', data_only=True)
points = workbook_values.sheetnames
values = workbook_values[points[0]]
with open('output2.csv', 'w', newline='') as csvfile:
    writer = csv.writer(
        csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL)
    for value in values:
        for elem in value:
            try:
                float(elem.value)
                f_value = float(elem.value)
                if f_value == elem.value:
                    lst.append(float(elem.value))
                else:
                    lst.append(int(elem.value))
            except ValueError:
                lst.append(elem.value)
            except TypeError:
                lst.append('')
    writer.writerow(lst)
